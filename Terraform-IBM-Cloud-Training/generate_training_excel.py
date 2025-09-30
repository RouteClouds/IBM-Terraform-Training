#!/usr/bin/env python3
"""
IBM Cloud Terraform Training Course Excel Generator

This script generates a professional Excel workbook with multiple worksheets
based on the IBM Cloud Terraform training content from Client-Reply.md.

Requirements:
- Python virtual environment at diagram-env/
- openpyxl library for Excel generation
- Client-Reply.md file in the project root

Output: IBM_Terraform_Training_Course_Details.xlsx
"""

import os
import sys
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime

class TerraformTrainingExcelGenerator:
    """Professional Excel workbook generator for IBM Cloud Terraform training course."""
    
    def __init__(self):
        """Initialize the Excel generator with professional styling."""
        self.workbook = Workbook()
        self.setup_styles()
        
    def setup_styles(self):
        """Define professional styling for the Excel workbook."""
        # Color scheme - IBM Blue theme
        self.colors = {
            'header_bg': 'C6E0FF',      # Light blue
            'subheader_bg': 'E6F3FF',   # Very light blue
            'accent_bg': 'F0F8FF',      # Alice blue
            'border_color': '4A90E2',   # IBM blue
            'text_dark': '1F2937',      # Dark gray
        }
        
        # Font styles
        self.fonts = {
            'title': Font(name='Calibri', size=16, bold=True, color='1F2937'),
            'header': Font(name='Calibri', size=12, bold=True, color='1F2937'),
            'subheader': Font(name='Calibri', size=11, bold=True, color='1F2937'),
            'normal': Font(name='Calibri', size=10, color='1F2937'),
            'small': Font(name='Calibri', size=9, color='1F2937'),
        }
        
        # Fill styles
        self.fills = {
            'header': PatternFill(start_color=self.colors['header_bg'], 
                                end_color=self.colors['header_bg'], fill_type='solid'),
            'subheader': PatternFill(start_color=self.colors['subheader_bg'], 
                                   end_color=self.colors['subheader_bg'], fill_type='solid'),
            'accent': PatternFill(start_color=self.colors['accent_bg'], 
                                end_color=self.colors['accent_bg'], fill_type='solid'),
        }
        
        # Border styles
        thin_border = Side(border_style='thin', color=self.colors['border_color'])
        self.borders = {
            'all': Border(left=thin_border, right=thin_border, 
                         top=thin_border, bottom=thin_border),
            'bottom': Border(bottom=thin_border),
        }
        
        # Alignment styles
        self.alignments = {
            'center': Alignment(horizontal='center', vertical='center', wrap_text=True),
            'left': Alignment(horizontal='left', vertical='center', wrap_text=True),
            'left_top': Alignment(horizontal='left', vertical='top', wrap_text=True),
        }

    def apply_cell_style(self, cell, style_type='normal', fill_type=None, border_type=None, alignment_type='left'):
        """Apply consistent styling to a cell."""
        cell.font = self.fonts[style_type]
        if fill_type:
            cell.fill = self.fills[fill_type]
        if border_type:
            cell.border = self.borders[border_type]
        cell.alignment = self.alignments[alignment_type]

    def create_merged_cell(self, ws, cell_range, value, style_type='normal', fill_type=None, border_type=None, alignment_type='left'):
        """Create a merged cell with proper styling."""
        # Get the top-left cell
        start_cell = ws[cell_range.split(':')[0]]
        start_cell.value = value
        self.apply_cell_style(start_cell, style_type, fill_type, border_type, alignment_type)
        ws.merge_cells(cell_range)

    def create_course_overview_sheet(self):
        """Create the Course Overview worksheet."""
        # Remove default sheet and create new one
        if 'Sheet' in self.workbook.sheetnames:
            self.workbook.remove(self.workbook['Sheet'])
        
        ws = self.workbook.create_sheet('Course Overview', 0)
        
        # Title
        self.create_merged_cell(ws, 'A1:D1', 'IBM Cloud For Terraform - Training Program',
                               'title', 'header', 'all', 'center')
        ws.row_dimensions[1].height = 25

        # Date and subject
        ws['A3'] = 'Date:'
        ws['B3'] = 'September 29, 2025'
        ws['A4'] = 'Subject:'
        ws['B4'] = 'Comprehensive 4-Day Training Program'

        for cell in [ws['A3'], ws['A4']]:
            self.apply_cell_style(cell, 'subheader')
        for cell in [ws['B3'], ws['B4']]:
            self.apply_cell_style(cell, 'normal')

        # Executive Summary
        self.create_merged_cell(ws, 'A6:D6', 'Executive Summary',
                               'header', 'subheader', 'all', 'center')

        summary_text = ("Comprehensive IBM Cloud For Terraform training program designed for "
                       "beginner-to-intermediate IT professionals, cloud engineers, and DevOps "
                       "practitioners. Delivers enterprise-grade content with hands-on laboratory "
                       "sessions and complete sandbox testing environment.")

        self.create_merged_cell(ws, 'A7:D9', summary_text, 'normal', None, None, 'left_top')
        ws.row_dimensions[7].height = 60

        # Course Details
        self.create_merged_cell(ws, 'A11:D11', 'Course Details',
                               'header', 'subheader', 'all', 'center')
        
        course_details = [
            ['Duration:', '4 Days (32 hours)'],
            ['Format:', 'Instructor-led with hands-on laboratories'],
            ['Delivery:', 'Blended learning (30% theory, 50% hands-on, 20% assessment)'],
            ['Target Audience:', 'IT professionals, Cloud engineers, DevOps practitioners'],
            ['Level:', 'Beginner to Intermediate'],
        ]
        
        row = 12
        for detail in course_details:
            ws[f'A{row}'] = detail[0]
            self.apply_cell_style(ws[f'A{row}'], 'subheader', border_type='all')
            self.create_merged_cell(ws, f'B{row}:D{row}', detail[1], 'normal', None, 'all')
            row += 1

        # Learning Outcomes
        self.create_merged_cell(ws, 'A18:D18', 'Key Learning Outcomes',
                               'header', 'subheader', 'all', 'center')

        outcomes = [
            'Design and implement Infrastructure as Code solutions using Terraform on IBM Cloud',
            'Configure and manage IBM Cloud resources through automated provisioning',
            'Apply enterprise best practices for modularization, state management, and security',
            'Integrate Terraform workflows with CI/CD pipelines and IBM Cloud Schematics',
            'Troubleshoot and optimize infrastructure deployments for cost and performance'
        ]

        row = 19
        for i, outcome in enumerate(outcomes, 1):
            ws[f'A{row}'] = f'{i}.'
            self.apply_cell_style(ws[f'A{row}'], 'normal', border_type='all', alignment_type='center')
            self.create_merged_cell(ws, f'B{row}:D{row}', outcome, 'normal', None, 'all')
            ws.row_dimensions[row].height = 30
            row += 1

        # Course Highlights
        self.create_merged_cell(ws, 'A25:D25', 'Course Highlights',
                               'header', 'subheader', 'all', 'center')

        highlights = [
            'Real-world scenarios with quantified business value and ROI calculations',
            'Comprehensive hands-on labs with actual IBM Cloud resource provisioning',
            'Enterprise-grade code examples following industry best practices',
            'Professional assessment framework with practical validation exercises',
            'Complete sandbox environment for safe learning and experimentation'
        ]

        row = 26
        for highlight in highlights:
            ws[f'A{row}'] = '•'
            self.apply_cell_style(ws[f'A{row}'], 'normal', border_type='all', alignment_type='center')
            self.create_merged_cell(ws, f'B{row}:D{row}', highlight, 'normal', None, 'all')
            ws.row_dimensions[row].height = 25
            row += 1
        
        # Set column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 25

    def create_daily_syllabus_sheet(self):
        """Create the Daily Syllabus worksheet."""
        ws = self.workbook.create_sheet('Daily Syllabus')

        # Title
        self.create_merged_cell(ws, 'A1:F1', 'Daily Syllabus - 4-Day Training Program',
                               'title', 'header', 'all', 'center')
        ws.row_dimensions[1].height = 25
        
        # Headers
        headers = ['Day', 'Session', 'Topic', 'Duration', 'Content', 'Lab Exercise']
        for i, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=i, value=header)
            self.apply_cell_style(cell, 'header', 'subheader', 'all', 'center')
        
        # Daily content data
        daily_content = [
            # Day 1
            ['Day 1\nFoundation\n& Setup', 'Morning\n(4 hours)', 'Topic 1: IaC Concepts &\nIBM Cloud Integration', '2 hours', 
             'Infrastructure as Code principles\nIBM Cloud advantages\nROI analysis frameworks', 'Lab 1: Manual vs Automated\ncomparison (90 min)'],
            ['', '', 'Topic 2: Terraform CLI &\nProvider Installation', '2 hours', 
             'Terraform CLI installation\nIBM Cloud Provider setup\nEnvironment optimization', 'Lab 2: Environment setup\nand validation (90 min)'],
            ['', 'Afternoon\n(4 hours)', 'Hands-on Practice', '4 hours', 
             'Environment validation\nTroubleshooting\nKnowledge reinforcement', 'Assessment: Foundation\nvalidation'],
            
            # Day 2
            ['Day 2\nCore Skills\nDevelopment', 'Morning\n(4 hours)', 'Topic 3: Core Terraform\nWorkflow', '2 hours', 
             'Project structure\nEssential commands\nProvider configuration', 'Lab 3: Complete workflow\nimplementation (120 min)'],
            ['', '', 'Topic 4: Resource\nProvisioning & Management', '2 hours', 
             'IBM Cloud resource definitions\nHCL syntax and variables\nResource dependencies', 'Lab 4: Multi-resource\ndeployment (120 min)'],
            ['', 'Afternoon\n(4 hours)', 'Advanced Techniques', '4 hours', 
             'Resource management\nError handling\nPerformance optimization', 'Assessment: Core skills\nevaluation'],
            
            # Day 3
            ['Day 3\nBest Practices\n& Advanced', 'Morning\n(4 hours)', 'Topic 5: Modularization\n& Best Practices', '2 hours', 
             'Reusable modules\nConfiguration organization\nVersion control with Git', 'Lab 5: Module development\n(120 min)'],
            ['', '', 'Topic 6: State Management', '2 hours', 
             'Local and remote state\nState locking\nDrift detection', 'Lab 6: Remote state\nconfiguration (120 min)'],
            ['', 'Afternoon\n(4 hours)', 'Advanced Scenarios', '4 hours', 
             'State management scenarios\nTeam collaboration\nDisaster recovery', 'Assessment: Best practices\nimplementation'],
            
            # Day 4
            ['Day 4\nSecurity &\nIntegration', 'Morning\n(4 hours)', 'Topic 7: Security\n& Compliance', '2 hours', 
             'Secrets management\nIAM integration\nSecurity scanning', 'Lab 7: Secure infrastructure\ndeployment (120 min)'],
            ['', '', 'Topic 8: Automation\n& Advanced Integration', '2 hours', 
             'CI/CD integration\nIBM Cloud Schematics\nTroubleshooting', 'Lab 8: Automation pipeline\n(120 min)'],
            ['', 'Afternoon\n(4 hours)', 'Enterprise Integration', '4 hours', 
             'Enterprise scenarios\nAdvanced troubleshooting\nCertification prep', 'Final Assessment:\nComprehensive evaluation'],
        ]
        
        # Populate data
        row = 4
        for content in daily_content:
            for col, value in enumerate(content, 1):
                cell = ws.cell(row=row, column=col, value=value)
                if col == 1 and value:  # Day column with content
                    self.apply_cell_style(cell, 'subheader', 'accent', 'all', 'center')
                elif col == 2:  # Session column
                    self.apply_cell_style(cell, 'subheader', border_type='all', alignment_type='center')
                else:
                    self.apply_cell_style(cell, 'normal', border_type='all', alignment_type='left_top')
                
                ws.row_dimensions[row].height = 45
            row += 1
        
        # Set column widths
        column_widths = [15, 12, 20, 10, 25, 20]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

    def create_lab_sessions_sheet(self):
        """Create the Lab Sessions worksheet."""
        ws = self.workbook.create_sheet('Lab Sessions')
        
        # Title
        ws['A1'] = 'Laboratory Sessions - Hands-On Learning'
        ws.merge_cells('A1:E1')
        self.apply_cell_style(ws['A1'], 'title', 'header', 'all', 'center')
        ws.row_dimensions[1].height = 25
        
        # Summary
        ws['A3'] = 'Lab Overview'
        ws.merge_cells('A3:E3')
        self.apply_cell_style(ws['A3'], 'header', 'subheader', 'all', 'center')
        
        overview_text = ("Over 14 hours of practical, hands-on experience with real IBM Cloud "
                        "resource provisioning. Each lab includes clear objectives, step-by-step "
                        "instructions, validation procedures, and troubleshooting guidance.")
        
        ws['A4'] = overview_text
        ws.merge_cells('A4:E5')
        self.apply_cell_style(ws['A4'], 'normal', alignment_type='left_top')
        ws.row_dimensions[4].height = 40
        
        # Lab details headers
        headers = ['Lab', 'Title', 'Duration', 'Objectives', 'Key Activities']
        for i, header in enumerate(headers, 1):
            cell = ws.cell(row=7, column=i, value=header)
            self.apply_cell_style(cell, 'header', 'subheader', 'all', 'center')
        
        # Lab data
        lab_data = [
            ['Lab 1', 'Infrastructure Comparison', '90 min', 
             'Compare manual vs automated provisioning', 
             'Manual setup, Terraform automation, Cost analysis, Version control demo'],
            ['Lab 2', 'Environment Setup', '90 min', 
             'Complete development environment configuration', 
             'CLI installation, Authentication setup, Connectivity validation'],
            ['Lab 3', 'Core Workflow', '120 min', 
             'End-to-end Terraform workflow implementation', 
             'VPC deployment, Resource lifecycle, Command execution'],
            ['Lab 4', 'Resource Management', '120 min', 
             'Complex multi-resource scenarios', 
             'Dependency management, Attribute references, Data sources'],
            ['Lab 5', 'Modularization', '120 min', 
             'Reusable module development', 
             'Module creation, Registry management, Team collaboration'],
            ['Lab 6', 'State Management', '120 min', 
             'Remote state configuration', 
             'Object Storage setup, State locking, Drift detection'],
            ['Lab 7', 'Security Implementation', '120 min', 
             'Secure credential management', 
             'IAM integration, Security scanning, Compliance validation'],
            ['Lab 8', 'Automation Pipeline', '120 min', 
             'Complete CI/CD integration', 
             'Schematics automation, Pipeline setup, Production deployment'],
        ]
        
        # Populate lab data
        row = 8
        for lab in lab_data:
            for col, value in enumerate(lab, 1):
                cell = ws.cell(row=row, column=col, value=value)
                if col == 1:  # Lab number
                    self.apply_cell_style(cell, 'subheader', 'accent', 'all', 'center')
                elif col == 2:  # Title
                    self.apply_cell_style(cell, 'subheader', border_type='all')
                else:
                    self.apply_cell_style(cell, 'normal', border_type='all', alignment_type='left_top')
                
                ws.row_dimensions[row].height = 50
            row += 1
        
        # Lab structure section
        ws[f'A{row+2}'] = 'Lab Structure Components'
        ws.merge_cells(f'A{row+2}:E{row+2}')
        self.apply_cell_style(ws[f'A{row+2}'], 'header', 'subheader', 'all', 'center')
        
        structure_items = [
            'Clear Learning Objectives - Specific, measurable outcomes',
            'Step-by-Step Instructions - Detailed procedures with validation checkpoints',
            'Real Resource Deployment - Actual IBM Cloud infrastructure provisioning',
            'Cost Estimation - Transparent pricing and resource optimization',
            'Validation Procedures - Technical verification of successful completion',
            'Troubleshooting Guidance - Common issues and resolution strategies',
            'Extension Activities - Advanced challenges for accelerated learners'
        ]
        
        start_row = row + 3
        for i, item in enumerate(structure_items):
            ws[f'A{start_row + i}'] = '•'
            ws[f'B{start_row + i}'] = item
            self.apply_cell_style(ws[f'A{start_row + i}'], 'normal', border_type='all', alignment_type='center')
            self.apply_cell_style(ws[f'B{start_row + i}'], 'normal', border_type='all')
            ws.merge_cells(f'B{start_row + i}:E{start_row + i}')
            ws.row_dimensions[start_row + i].height = 25
        
        # Set column widths
        column_widths = [8, 20, 12, 25, 30]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

    def create_sandbox_environment_sheet(self):
        """Create the Sandbox Environment worksheet."""
        ws = self.workbook.create_sheet('Sandbox Environment')

        # Title
        ws['A1'] = 'Sandbox Environment - Complete Learning Infrastructure'
        ws.merge_cells('A1:D1')
        self.apply_cell_style(ws['A1'], 'title', 'header', 'all', 'center')
        ws.row_dimensions[1].height = 25

        # Overview
        ws['A3'] = 'Environment Overview'
        ws.merge_cells('A3:D3')
        self.apply_cell_style(ws['A3'], 'header', 'subheader', 'all', 'center')

        overview_text = ("Comprehensive sandbox environment ensuring safe, cost-controlled learning "
                        "without impacting production systems. Includes dedicated IBM Cloud resources, "
                        "development tools, and automated management.")

        ws['A4'] = overview_text
        ws.merge_cells('A4:D5')
        self.apply_cell_style(ws['A4'], 'normal', alignment_type='left_top')
        ws.row_dimensions[4].height = 40

        # Technical Infrastructure
        ws['A7'] = 'Technical Infrastructure'
        ws.merge_cells('A7:D7')
        self.apply_cell_style(ws['A7'], 'header', 'subheader', 'all', 'center')

        infrastructure_items = [
            ['Component', 'Description', 'Benefits'],
            ['Dedicated IBM Cloud Resources', 'Isolated training environment with resource groups', 'Safe learning without production impact'],
            ['Cost Controls', 'Automated spending limits and resource quotas', 'Predictable training costs'],
            ['Security Isolation', 'No access to production environments', 'Risk-free experimentation'],
            ['Automated Cleanup', 'Resources automatically removed after training', 'No ongoing costs'],
        ]

        row = 8
        for i, item in enumerate(infrastructure_items):
            for col, value in enumerate(item, 1):
                cell = ws.cell(row=row, column=col, value=value)
                if i == 0:  # Header row
                    self.apply_cell_style(cell, 'subheader', 'accent', 'all', 'center')
                else:
                    self.apply_cell_style(cell, 'normal', border_type='all', alignment_type='left_top')

                if col == 3:  # Merge last column
                    ws.merge_cells(f'{get_column_letter(col)}{row}:D{row}')

                ws.row_dimensions[row].height = 30
            row += 1

        # Development Tools
        ws['A14'] = 'Development Tools'
        ws.merge_cells('A14:D14')
        self.apply_cell_style(ws['A14'], 'header', 'subheader', 'all', 'center')

        tools = [
            'Terraform CLI (v1.5.0+) - Latest version with IBM Cloud provider',
            'IBM Cloud CLI - Complete toolchain with required plugins',
            'VS Code with Extensions - Terraform syntax highlighting and validation',
            'Git Integration - Version control for collaboration exercises',
            'Development Environment - Pre-configured workstations'
        ]

        row = 15
        for tool in tools:
            ws[f'A{row}'] = '•'
            ws[f'B{row}'] = tool
            self.apply_cell_style(ws[f'A{row}'], 'normal', border_type='all', alignment_type='center')
            self.apply_cell_style(ws[f'B{row}'], 'normal', border_type='all')
            ws.merge_cells(f'B{row}:D{row}')
            ws.row_dimensions[row].height = 25
            row += 1

        # Setup Process
        ws['A21'] = 'Setup Process'
        ws.merge_cells('A21:D21')
        self.apply_cell_style(ws['A21'], 'header', 'subheader', 'all', 'center')

        setup_phases = [
            ['Phase', 'Activities', 'Timeline'],
            ['Pre-Training', 'Account provisioning, Access configuration, Environment validation', 'Before training'],
            ['During Training', 'Guided setup, Validation checkpoints, Ongoing support', 'Day 1-4'],
            ['Post-Training', 'Resource cleanup, Knowledge transfer, Continued access options', 'After training'],
        ]

        row = 22
        for i, phase in enumerate(setup_phases):
            for col, value in enumerate(phase, 1):
                cell = ws.cell(row=row, column=col, value=value)
                if i == 0:  # Header row
                    self.apply_cell_style(cell, 'subheader', 'accent', 'all', 'center')
                else:
                    self.apply_cell_style(cell, 'normal', border_type='all', alignment_type='left_top')

                if col == 2:  # Merge activities column
                    ws.merge_cells(f'B{row}:C{row}')

                ws.row_dimensions[row].height = 35
            row += 1

        # Set column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20

    def create_learning_objectives_sheet(self):
        """Create the Learning Objectives worksheet."""
        ws = self.workbook.create_sheet('Learning Objectives')

        # Title
        ws['A1'] = 'Learning Objectives & Business Benefits'
        ws.merge_cells('A1:D1')
        self.apply_cell_style(ws['A1'], 'title', 'header', 'all', 'center')
        ws.row_dimensions[1].height = 25

        # Skills Development
        ws['A3'] = 'Immediate Skills Development'
        ws.merge_cells('A3:D3')
        self.apply_cell_style(ws['A3'], 'header', 'subheader', 'all', 'center')

        skills = [
            ['Skill Area', 'Description', 'Practical Application'],
            ['Practical Expertise', 'Hands-on experience with real IBM Cloud deployments', 'Immediate infrastructure automation capability'],
            ['Industry Best Practices', 'Enterprise-grade patterns and methodologies', 'Professional-level implementation standards'],
            ['Cost Optimization', 'Proven strategies for resource efficiency', '30-50% infrastructure cost savings'],
            ['Security Implementation', 'Comprehensive security and compliance practices', 'Enterprise-grade security controls'],
        ]

        row = 4
        for i, skill in enumerate(skills):
            for col, value in enumerate(skill, 1):
                cell = ws.cell(row=row, column=col, value=value)
                if i == 0:  # Header row
                    self.apply_cell_style(cell, 'subheader', 'accent', 'all', 'center')
                else:
                    self.apply_cell_style(cell, 'normal', border_type='all', alignment_type='left_top')

                ws.row_dimensions[row].height = 35
            row += 1

        # Business Value
        ws['A10'] = 'Business Value Delivered'
        ws.merge_cells('A10:D10')
        self.apply_cell_style(ws['A10'], 'header', 'subheader', 'all', 'center')

        business_values = [
            ['Metric', 'Improvement', 'Impact'],
            ['Infrastructure Costs', '30-50% reduction', 'Significant cost savings through automation'],
            ['Deployment Speed', '70% faster provisioning', 'Rapid infrastructure delivery'],
            ['Configuration Errors', '90% reduction', 'Improved reliability and uptime'],
            ['Compliance', 'Automated governance', 'Enhanced security and audit readiness'],
        ]

        row = 11
        for i, value in enumerate(business_values):
            for col, item in enumerate(value, 1):
                cell = ws.cell(row=row, column=col, value=item)
                if i == 0:  # Header row
                    self.apply_cell_style(cell, 'subheader', 'accent', 'all', 'center')
                else:
                    self.apply_cell_style(cell, 'normal', border_type='all', alignment_type='left_top')

                ws.row_dimensions[row].height = 30
            row += 1

        # Career Advancement
        ws['A17'] = 'Career Advancement Opportunities'
        ws.merge_cells('A17:D17')
        self.apply_cell_style(ws['A17'], 'header', 'subheader', 'all', 'center')

        career_benefits = [
            'High-Demand Skills - Infrastructure as Code expertise in growing demand',
            'IBM Cloud Certification - Preparation for professional certifications',
            'Enterprise Readiness - Skills applicable to large-scale deployments',
            'Competitive Advantage - Specialized IBM Cloud automation knowledge',
            'Leadership Opportunities - Technical expertise for team leadership roles'
        ]

        row = 18
        for benefit in career_benefits:
            ws[f'A{row}'] = '•'
            ws[f'B{row}'] = benefit
            self.apply_cell_style(ws[f'A{row}'], 'normal', border_type='all', alignment_type='center')
            self.apply_cell_style(ws[f'B{row}'], 'normal', border_type='all')
            ws.merge_cells(f'B{row}:D{row}')
            ws.row_dimensions[row].height = 25
            row += 1

        # Set column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 25

    def create_delivery_options_sheet(self):
        """Create the Delivery Options worksheet."""
        ws = self.workbook.create_sheet('Delivery Options')

        # Title
        ws['A1'] = 'Training Delivery Options'
        ws.merge_cells('A1:E1')
        self.apply_cell_style(ws['A1'], 'title', 'header', 'all', 'center')
        ws.row_dimensions[1].height = 25

        # Options comparison
        ws['A3'] = 'Delivery Format Comparison'
        ws.merge_cells('A3:E3')
        self.apply_cell_style(ws['A3'], 'header', 'subheader', 'all', 'center')

        # Headers
        headers = ['Option', 'Timeline', 'Content Coverage', 'Benefits', 'Ideal For']
        for i, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=i, value=header)
            self.apply_cell_style(cell, 'subheader', 'accent', 'all', 'center')

        # Delivery options data
        options_data = [
            ['Complete 4-Day Program\n(Recommended)', 'Available within\n3-4 weeks', 'All 8 topics with\ncomprehensive coverage',
             'Complete skill development\nEnterprise-grade expertise\nFull certification prep',
             'Organizations seeking\ncomprehensive IBM Cloud\nTerraform expertise'],
            ['Accelerated 3-Day Program', 'Available\nimmediately', 'Topics 1-6\n(Foundation through\nBest Practices)',
             'Rapid skill development\nCore competencies\nImmediate value',
             'Teams needing immediate\nTerraform capabilities'],
            ['Phased Delivery', 'Phase 1: Immediate\nPhase 2: 3-4 weeks', 'Phase 1: Topics 1-6\nPhase 2: Topics 7-8',
             'Immediate value\nProgressive enhancement\nFlexible scheduling',
             'Organizations with urgent\nneeds and long-term\ndevelopment goals'],
        ]

        # Populate options data
        row = 5
        for option in options_data:
            for col, value in enumerate(option, 1):
                cell = ws.cell(row=row, column=col, value=value)
                if col == 1:  # Option name
                    self.apply_cell_style(cell, 'subheader', 'subheader', 'all', 'center')
                else:
                    self.apply_cell_style(cell, 'normal', border_type='all', alignment_type='left_top')

                ws.row_dimensions[row].height = 60
            row += 1

        # Recommendation section
        ws['A9'] = 'Recommendation'
        ws.merge_cells('A9:E9')
        self.apply_cell_style(ws['A9'], 'header', 'subheader', 'all', 'center')

        recommendation_text = ("We recommend the Complete 4-Day Program for maximum value and comprehensive "
                              "skill development. However, the Accelerated 3-Day Program is available for "
                              "immediate delivery if urgent training needs exist.")

        ws['A10'] = recommendation_text
        ws.merge_cells('A10:E11')
        self.apply_cell_style(ws['A10'], 'normal', 'accent', 'all', 'left_top')
        ws.row_dimensions[10].height = 40

        # Set column widths
        column_widths = [20, 15, 20, 25, 25]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

    def create_prerequisites_sheet(self):
        """Create the Prerequisites worksheet."""
        ws = self.workbook.create_sheet('Prerequisites')

        # Title
        ws['A1'] = 'Course Prerequisites & Requirements'
        ws.merge_cells('A1:D1')
        self.apply_cell_style(ws['A1'], 'title', 'header', 'all', 'center')
        ws.row_dimensions[1].height = 25

        # Required Knowledge
        ws['A3'] = 'Required Knowledge'
        ws.merge_cells('A3:D3')
        self.apply_cell_style(ws['A3'], 'header', 'subheader', 'all', 'center')

        required_knowledge = [
            ['Area', 'Requirement', 'Description'],
            ['Cloud Computing Basics', 'Essential', 'Understanding of cloud service models and deployment'],
            ['Command Line Proficiency', 'Essential', 'Comfortable with terminal/command prompt operations'],
            ['Infrastructure Fundamentals', 'Essential', 'Basic networking, compute, and storage concepts'],
        ]

        row = 4
        for i, knowledge in enumerate(required_knowledge):
            for col, value in enumerate(knowledge, 1):
                cell = ws.cell(row=row, column=col, value=value)
                if i == 0:  # Header row
                    self.apply_cell_style(cell, 'subheader', 'accent', 'all', 'center')
                else:
                    self.apply_cell_style(cell, 'normal', border_type='all', alignment_type='left_top')

                if col == 3:  # Merge description column
                    ws.merge_cells(f'C{row}:D{row}')

                ws.row_dimensions[row].height = 30
            row += 1

        # Recommended Experience
        ws['A9'] = 'Recommended Experience'
        ws.merge_cells('A9:D9')
        self.apply_cell_style(ws['A9'], 'header', 'subheader', 'all', 'center')

        recommended_exp = [
            'Any Cloud Platform - Previous experience with AWS, Azure, or IBM Cloud helpful',
            'Configuration Management - Familiarity with automation tools beneficial',
            'Development Practices - Basic understanding of version control (Git) preferred',
            'Infrastructure Operations - Experience with server and network management useful'
        ]

        row = 10
        for exp in recommended_exp:
            ws[f'A{row}'] = '•'
            ws[f'B{row}'] = exp
            self.apply_cell_style(ws[f'A{row}'], 'normal', border_type='all', alignment_type='center')
            self.apply_cell_style(ws[f'B{row}'], 'normal', border_type='all')
            ws.merge_cells(f'B{row}:D{row}')
            ws.row_dimensions[row].height = 25
            row += 1

        # Technical Requirements
        ws['A15'] = 'Technical Requirements'
        ws.merge_cells('A15:D15')
        self.apply_cell_style(ws['A15'], 'header', 'subheader', 'all', 'center')

        tech_requirements = [
            ['Requirement', 'Specification', 'Notes'],
            ['Laptop/Workstation', 'Modern computer with 8GB+ RAM', 'Internet connectivity required'],
            ['Web Browser', 'Chrome, Firefox, or Safari (current version)', 'For IBM Cloud console access'],
            ['SSH Client', 'Terminal or PuTTY for secure access', 'Provided if needed'],
            ['IBM Cloud Account', 'Training account provided', 'Or existing account with appropriate access'],
        ]

        row = 16
        for i, req in enumerate(tech_requirements):
            for col, value in enumerate(req, 1):
                cell = ws.cell(row=row, column=col, value=value)
                if i == 0:  # Header row
                    self.apply_cell_style(cell, 'subheader', 'accent', 'all', 'center')
                else:
                    self.apply_cell_style(cell, 'normal', border_type='all', alignment_type='left_top')

                ws.row_dimensions[row].height = 30
            row += 1

        # Set column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20

    def generate_excel_file(self, output_filename='IBM_Terraform_Training_Course_Details.xlsx'):
        """Generate the complete Excel workbook."""
        try:
            print("Generating IBM Cloud Terraform Training Excel workbook...")

            # Create all worksheets
            self.create_course_overview_sheet()
            self.create_daily_syllabus_sheet()
            self.create_lab_sessions_sheet()
            self.create_sandbox_environment_sheet()
            self.create_learning_objectives_sheet()
            self.create_delivery_options_sheet()
            self.create_prerequisites_sheet()

            # Save the workbook
            output_path = Path(output_filename)
            self.workbook.save(output_path)

            print(f"✅ Excel workbook successfully generated: {output_path.absolute()}")
            print(f"📊 Worksheets created: {len(self.workbook.sheetnames)}")
            print(f"📋 Worksheet names: {', '.join(self.workbook.sheetnames)}")

            return True

        except Exception as e:
            print(f"❌ Error generating Excel workbook: {str(e)}")
            return False

def main():
    """Main function to generate the Excel workbook."""
    print("🚀 Starting IBM Cloud Terraform Training Excel Generator")
    print("=" * 60)
    
    # Check if Client-Reply.md exists
    client_reply_path = Path('../Client-Reply.md')
    if not client_reply_path.exists():
        print(f"⚠️  Warning: Client-Reply.md not found at {client_reply_path.absolute()}")
        print("   Proceeding with embedded content...")
    
    # Generate the Excel workbook
    generator = TerraformTrainingExcelGenerator()
    success = generator.generate_excel_file()
    
    if success:
        print("\n🎉 Excel generation completed successfully!")
        print("📄 The file is ready for client presentation.")
    else:
        print("\n❌ Excel generation failed. Please check the error messages above.")
        sys.exit(1)

if __name__ == "__main__":
    main()

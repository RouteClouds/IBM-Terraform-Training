#!/usr/bin/env python3
"""
IBM Cloud Terraform Training Course Excel Generator - Simplified Version

This script generates a professional Excel workbook with multiple worksheets
based on the IBM Cloud Terraform training content from Client-Reply.md.

Requirements:
- Python virtual environment at diagram-env/
- openpyxl library for Excel generation

Output: IBM_Terraform_Training_Course_Details.xlsx
"""

import os
import sys
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime

class TerraformTrainingExcelGenerator:
    """Professional Excel workbook generator for IBM Cloud Terraform training course."""
    
    def __init__(self):
        """Initialize the Excel generator with professional styling."""
        self.workbook = Workbook()
        self.setup_styles()
        
    def setup_styles(self):
        """Define professional styling for the Excel workbook."""
        # Color scheme - IBM Blue theme
        self.colors = {
            'header_bg': 'C6E0FF',      # Light blue
            'subheader_bg': 'E6F3FF',   # Very light blue
            'accent_bg': 'F0F8FF',      # Alice blue
            'border_color': '4A90E2',   # IBM blue
        }
        
        # Font styles
        self.fonts = {
            'title': Font(name='Calibri', size=16, bold=True, color='1F2937'),
            'header': Font(name='Calibri', size=12, bold=True, color='1F2937'),
            'subheader': Font(name='Calibri', size=11, bold=True, color='1F2937'),
            'normal': Font(name='Calibri', size=10, color='1F2937'),
        }
        
        # Fill styles
        self.fills = {
            'header': PatternFill(start_color=self.colors['header_bg'], 
                                end_color=self.colors['header_bg'], fill_type='solid'),
            'subheader': PatternFill(start_color=self.colors['subheader_bg'], 
                                   end_color=self.colors['subheader_bg'], fill_type='solid'),
            'accent': PatternFill(start_color=self.colors['accent_bg'], 
                                end_color=self.colors['accent_bg'], fill_type='solid'),
        }
        
        # Border styles
        thin_border = Side(border_style='thin', color=self.colors['border_color'])
        self.borders = {
            'all': Border(left=thin_border, right=thin_border, 
                         top=thin_border, bottom=thin_border),
        }
        
        # Alignment styles
        self.alignments = {
            'center': Alignment(horizontal='center', vertical='center', wrap_text=True),
            'left': Alignment(horizontal='left', vertical='center', wrap_text=True),
            'left_top': Alignment(horizontal='left', vertical='top', wrap_text=True),
        }

    def apply_cell_style(self, cell, style_type='normal', fill_type=None, border_type=None, alignment_type='left'):
        """Apply consistent styling to a cell."""
        cell.font = self.fonts[style_type]
        if fill_type:
            cell.fill = self.fills[fill_type]
        if border_type:
            cell.border = self.borders[border_type]
        cell.alignment = self.alignments[alignment_type]

    def create_course_overview_sheet(self):
        """Create the Course Overview worksheet."""
        # Remove default sheet and create new one
        if 'Sheet' in self.workbook.sheetnames:
            self.workbook.remove(self.workbook['Sheet'])
        
        ws = self.workbook.create_sheet('Course Overview', 0)
        
        # Title
        ws['A1'] = 'IBM Cloud For Terraform - Training Program'
        self.apply_cell_style(ws['A1'], 'title', 'header', 'all', 'center')
        ws.row_dimensions[1].height = 25
        
        # Course Information
        course_info = [
            ['', 'Course Information', '', ''],
            ['Duration:', '4 Days (32 hours)', '', ''],
            ['Format:', 'Instructor-led with hands-on laboratories', '', ''],
            ['Delivery:', 'Blended learning (30% theory, 50% hands-on, 20% assessment)', '', ''],
            ['Target Audience:', 'IT professionals, Cloud engineers, DevOps practitioners', '', ''],
            ['Level:', 'Beginner to Intermediate', '', ''],
            ['', '', '', ''],
            ['', 'Learning Outcomes', '', ''],
            ['1.', 'Design and implement Infrastructure as Code solutions using Terraform on IBM Cloud', '', ''],
            ['2.', 'Configure and manage IBM Cloud resources through automated provisioning', '', ''],
            ['3.', 'Apply enterprise best practices for modularization, state management, and security', '', ''],
            ['4.', 'Integrate Terraform workflows with CI/CD pipelines and IBM Cloud Schematics', '', ''],
            ['5.', 'Troubleshoot and optimize infrastructure deployments for cost and performance', '', ''],
            ['', '', '', ''],
            ['', 'Course Highlights', '', ''],
            ['•', 'Real-world scenarios with quantified business value and ROI calculations', '', ''],
            ['•', 'Comprehensive hands-on labs with actual IBM Cloud resource provisioning', '', ''],
            ['•', 'Enterprise-grade code examples following industry best practices', '', ''],
            ['•', 'Professional assessment framework with practical validation exercises', '', ''],
            ['•', 'Complete sandbox environment for safe learning and experimentation', '', ''],
        ]
        
        # Populate course information
        for row_idx, row_data in enumerate(course_info, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                
                if 'Information' in str(value) or 'Learning Outcomes' in str(value) or 'Course Highlights' in str(value):
                    self.apply_cell_style(cell, 'header', 'subheader', 'all', 'center')
                elif value and value.endswith(':'):
                    self.apply_cell_style(cell, 'subheader', border_type='all')
                elif value and (value.startswith(('1.', '2.', '3.', '4.', '5.')) or value == '•'):
                    self.apply_cell_style(cell, 'normal', border_type='all', alignment_type='center')
                elif value:
                    self.apply_cell_style(cell, 'normal', border_type='all')
                
                if col_idx == 2 and value:  # Main content column
                    ws.row_dimensions[row_idx].height = 25
        
        # Set column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 60
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15

    def create_daily_syllabus_sheet(self):
        """Create the Daily Syllabus worksheet."""
        ws = self.workbook.create_sheet('Daily Syllabus')
        
        # Title
        ws['A1'] = 'Daily Syllabus - 4-Day Training Program'
        self.apply_cell_style(ws['A1'], 'title', 'header', 'all', 'center')
        ws.row_dimensions[1].height = 25
        
        # Headers
        headers = ['Day', 'Session', 'Topic', 'Duration', 'Content', 'Lab Exercise']
        for i, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=i, value=header)
            self.apply_cell_style(cell, 'header', 'subheader', 'all', 'center')
        
        # Daily content data
        daily_content = [
            ['Day 1: Foundation & Setup', 'Morning (4 hours)', 'Topic 1: IaC Concepts & IBM Cloud Integration', '2 hours', 
             'Infrastructure as Code principles, IBM Cloud advantages, ROI analysis frameworks', 'Lab 1: Manual vs Automated comparison (90 min)'],
            ['', '', 'Topic 2: Terraform CLI & Provider Installation', '2 hours', 
             'Terraform CLI installation, IBM Cloud Provider setup, Environment optimization', 'Lab 2: Environment setup and validation (90 min)'],
            ['', 'Afternoon (4 hours)', 'Hands-on Practice', '4 hours', 
             'Environment validation, Troubleshooting, Knowledge reinforcement', 'Assessment: Foundation validation'],
            
            ['Day 2: Core Skills Development', 'Morning (4 hours)', 'Topic 3: Core Terraform Workflow', '2 hours', 
             'Project structure, Essential commands, Provider configuration', 'Lab 3: Complete workflow implementation (120 min)'],
            ['', '', 'Topic 4: Resource Provisioning & Management', '2 hours', 
             'IBM Cloud resource definitions, HCL syntax and variables, Resource dependencies', 'Lab 4: Multi-resource deployment (120 min)'],
            ['', 'Afternoon (4 hours)', 'Advanced Techniques', '4 hours', 
             'Resource management, Error handling, Performance optimization', 'Assessment: Core skills evaluation'],
            
            ['Day 3: Best Practices & Advanced', 'Morning (4 hours)', 'Topic 5: Modularization & Best Practices', '2 hours', 
             'Reusable modules, Configuration organization, Version control with Git', 'Lab 5: Module development (120 min)'],
            ['', '', 'Topic 6: State Management', '2 hours', 
             'Local and remote state, State locking, Drift detection', 'Lab 6: Remote state configuration (120 min)'],
            ['', 'Afternoon (4 hours)', 'Advanced Scenarios', '4 hours', 
             'State management scenarios, Team collaboration, Disaster recovery', 'Assessment: Best practices implementation'],
            
            ['Day 4: Security & Integration', 'Morning (4 hours)', 'Topic 7: Security & Compliance', '2 hours', 
             'Secrets management, IAM integration, Security scanning', 'Lab 7: Secure infrastructure deployment (120 min)'],
            ['', '', 'Topic 8: Automation & Advanced Integration', '2 hours', 
             'CI/CD integration, IBM Cloud Schematics, Troubleshooting', 'Lab 8: Automation pipeline (120 min)'],
            ['', 'Afternoon (4 hours)', 'Enterprise Integration', '4 hours', 
             'Enterprise scenarios, Advanced troubleshooting, Certification prep', 'Final Assessment: Comprehensive evaluation'],
        ]
        
        # Populate data
        row = 4
        for content in daily_content:
            for col, value in enumerate(content, 1):
                cell = ws.cell(row=row, column=col, value=value)
                if col == 1 and value:  # Day column with content
                    self.apply_cell_style(cell, 'subheader', 'accent', 'all', 'center')
                elif col == 2:  # Session column
                    self.apply_cell_style(cell, 'subheader', border_type='all', alignment_type='center')
                else:
                    self.apply_cell_style(cell, 'normal', border_type='all', alignment_type='left_top')
                
                ws.row_dimensions[row].height = 45
            row += 1
        
        # Set column widths
        column_widths = [25, 15, 25, 12, 35, 25]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

    def create_lab_sessions_sheet(self):
        """Create the Lab Sessions worksheet."""
        ws = self.workbook.create_sheet('Lab Sessions')
        
        # Title
        ws['A1'] = 'Laboratory Sessions - Hands-On Learning'
        self.apply_cell_style(ws['A1'], 'title', 'header', 'all', 'center')
        ws.row_dimensions[1].height = 25
        
        # Lab details headers
        headers = ['Lab', 'Title', 'Duration', 'Objectives', 'Key Activities']
        for i, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=i, value=header)
            self.apply_cell_style(cell, 'header', 'subheader', 'all', 'center')
        
        # Lab data
        lab_data = [
            ['Lab 1', 'Infrastructure Comparison', '90 min', 
             'Compare manual vs automated provisioning', 
             'Manual setup, Terraform automation, Cost analysis, Version control demo'],
            ['Lab 2', 'Environment Setup', '90 min', 
             'Complete development environment configuration', 
             'CLI installation, Authentication setup, Connectivity validation'],
            ['Lab 3', 'Core Workflow', '120 min', 
             'End-to-end Terraform workflow implementation', 
             'VPC deployment, Resource lifecycle, Command execution'],
            ['Lab 4', 'Resource Management', '120 min', 
             'Complex multi-resource scenarios', 
             'Dependency management, Attribute references, Data sources'],
            ['Lab 5', 'Modularization', '120 min', 
             'Reusable module development', 
             'Module creation, Registry management, Team collaboration'],
            ['Lab 6', 'State Management', '120 min', 
             'Remote state configuration', 
             'Object Storage setup, State locking, Drift detection'],
            ['Lab 7', 'Security Implementation', '120 min', 
             'Secure credential management', 
             'IAM integration, Security scanning, Compliance validation'],
            ['Lab 8', 'Automation Pipeline', '120 min', 
             'Complete CI/CD integration', 
             'Schematics automation, Pipeline setup, Production deployment'],
        ]
        
        # Populate lab data
        row = 4
        for lab in lab_data:
            for col, value in enumerate(lab, 1):
                cell = ws.cell(row=row, column=col, value=value)
                if col == 1:  # Lab number
                    self.apply_cell_style(cell, 'subheader', 'accent', 'all', 'center')
                elif col == 2:  # Title
                    self.apply_cell_style(cell, 'subheader', border_type='all')
                else:
                    self.apply_cell_style(cell, 'normal', border_type='all', alignment_type='left_top')
                
                ws.row_dimensions[row].height = 50
            row += 1
        
        # Set column widths
        column_widths = [10, 25, 12, 30, 35]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

    def create_sandbox_environment_sheet(self):
        """Create the Sandbox Environment worksheet."""
        ws = self.workbook.create_sheet('Sandbox Environment')

        # Title
        ws['A1'] = 'Sandbox Environment - Complete Learning Infrastructure'
        self.apply_cell_style(ws['A1'], 'title', 'header', 'all', 'center')
        ws.row_dimensions[1].height = 25

        # Environment data
        environment_data = [
            ['', 'Environment Overview', ''],
            ['Description:', 'Comprehensive sandbox environment ensuring safe, cost-controlled learning without impacting production systems', ''],
            ['', '', ''],
            ['', 'Technical Infrastructure', ''],
            ['Component', 'Description', 'Benefits'],
            ['Dedicated IBM Cloud Resources', 'Isolated training environment with resource groups', 'Safe learning without production impact'],
            ['Cost Controls', 'Automated spending limits and resource quotas', 'Predictable training costs'],
            ['Security Isolation', 'No access to production environments', 'Risk-free experimentation'],
            ['Automated Cleanup', 'Resources automatically removed after training', 'No ongoing costs'],
            ['', '', ''],
            ['', 'Development Tools', ''],
            ['•', 'Terraform CLI (v1.5.0+) - Latest version with IBM Cloud provider', ''],
            ['•', 'IBM Cloud CLI - Complete toolchain with required plugins', ''],
            ['•', 'VS Code with Extensions - Terraform syntax highlighting and validation', ''],
            ['•', 'Git Integration - Version control for collaboration exercises', ''],
            ['•', 'Development Environment - Pre-configured workstations', ''],
            ['', '', ''],
            ['', 'Setup Process', ''],
            ['Phase', 'Activities', 'Timeline'],
            ['Pre-Training', 'Account provisioning, Access configuration, Environment validation', 'Before training'],
            ['During Training', 'Guided setup, Validation checkpoints, Ongoing support', 'Day 1-4'],
            ['Post-Training', 'Resource cleanup, Knowledge transfer, Continued access options', 'After training'],
        ]

        # Populate environment data
        for row_idx, row_data in enumerate(environment_data, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)

                if 'Overview' in str(value) or 'Infrastructure' in str(value) or 'Development Tools' in str(value) or 'Setup Process' in str(value):
                    self.apply_cell_style(cell, 'header', 'subheader', 'all', 'center')
                elif value and (value.endswith(':') or value in ['Component', 'Phase']):
                    self.apply_cell_style(cell, 'subheader', 'accent', 'all')
                elif value and value == '•':
                    self.apply_cell_style(cell, 'normal', border_type='all', alignment_type='center')
                elif value:
                    self.apply_cell_style(cell, 'normal', border_type='all', alignment_type='left_top')

                ws.row_dimensions[row_idx].height = 25

        # Set column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 25

    def create_delivery_options_sheet(self):
        """Create the Delivery Options worksheet."""
        ws = self.workbook.create_sheet('Delivery Options')

        # Title
        ws['A1'] = 'Training Delivery Options'
        self.apply_cell_style(ws['A1'], 'title', 'header', 'all', 'center')
        ws.row_dimensions[1].height = 25

        # Options data
        options_data = [
            ['', 'Delivery Format Comparison', '', '', ''],
            ['Option', 'Timeline', 'Content Coverage', 'Benefits', 'Ideal For'],
            ['Complete 4-Day Program (Recommended)', 'Available within 3-4 weeks', 'All 8 topics with comprehensive coverage',
             'Complete skill development, Enterprise-grade expertise, Full certification prep',
             'Organizations seeking comprehensive IBM Cloud Terraform expertise'],
            ['Accelerated 3-Day Program', 'Available immediately', 'Topics 1-6 (Foundation through Best Practices)',
             'Rapid skill development, Core competencies, Immediate value',
             'Teams needing immediate Terraform capabilities'],
            ['Phased Delivery', 'Phase 1: Immediate, Phase 2: 3-4 weeks', 'Phase 1: Topics 1-6, Phase 2: Topics 7-8',
             'Immediate value, Progressive enhancement, Flexible scheduling',
             'Organizations with urgent needs and long-term development goals'],
            ['', '', '', '', ''],
            ['', 'Recommendation', '', '', ''],
            ['', 'We recommend the Complete 4-Day Program for maximum value and comprehensive skill development. However, the Accelerated 3-Day Program is available for immediate delivery if urgent training needs exist.', '', '', ''],
        ]

        # Populate options data
        for row_idx, row_data in enumerate(options_data, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)

                if 'Comparison' in str(value) or 'Recommendation' in str(value):
                    self.apply_cell_style(cell, 'header', 'subheader', 'all', 'center')
                elif value in ['Option', 'Timeline', 'Content Coverage', 'Benefits', 'Ideal For']:
                    self.apply_cell_style(cell, 'subheader', 'accent', 'all', 'center')
                elif value and 'Program' in value:
                    self.apply_cell_style(cell, 'subheader', border_type='all', alignment_type='center')
                elif value:
                    self.apply_cell_style(cell, 'normal', border_type='all', alignment_type='left_top')

                ws.row_dimensions[row_idx].height = 40

        # Set column widths
        column_widths = [25, 20, 25, 30, 30]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

    def create_prerequisites_sheet(self):
        """Create the Prerequisites worksheet."""
        ws = self.workbook.create_sheet('Prerequisites')

        # Title
        ws['A1'] = 'Course Prerequisites & Requirements'
        self.apply_cell_style(ws['A1'], 'title', 'header', 'all', 'center')
        ws.row_dimensions[1].height = 25

        # Prerequisites data
        prereq_data = [
            ['', 'Required Knowledge', ''],
            ['Area', 'Requirement', 'Description'],
            ['Cloud Computing Basics', 'Essential', 'Understanding of cloud service models and deployment'],
            ['Command Line Proficiency', 'Essential', 'Comfortable with terminal/command prompt operations'],
            ['Infrastructure Fundamentals', 'Essential', 'Basic networking, compute, and storage concepts'],
            ['', '', ''],
            ['', 'Recommended Experience', ''],
            ['•', 'Any Cloud Platform - Previous experience with AWS, Azure, or IBM Cloud helpful', ''],
            ['•', 'Configuration Management - Familiarity with automation tools beneficial', ''],
            ['•', 'Development Practices - Basic understanding of version control (Git) preferred', ''],
            ['•', 'Infrastructure Operations - Experience with server and network management useful', ''],
            ['', '', ''],
            ['', 'Technical Requirements', ''],
            ['Requirement', 'Specification', 'Notes'],
            ['Laptop/Workstation', 'Modern computer with 8GB+ RAM', 'Internet connectivity required'],
            ['Web Browser', 'Chrome, Firefox, or Safari (current version)', 'For IBM Cloud console access'],
            ['SSH Client', 'Terminal or PuTTY for secure access', 'Provided if needed'],
            ['IBM Cloud Account', 'Training account provided', 'Or existing account with appropriate access'],
        ]

        # Populate prerequisites data
        for row_idx, row_data in enumerate(prereq_data, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)

                if 'Knowledge' in str(value) or 'Experience' in str(value) or 'Requirements' in str(value):
                    self.apply_cell_style(cell, 'header', 'subheader', 'all', 'center')
                elif value in ['Area', 'Requirement', 'Specification']:
                    self.apply_cell_style(cell, 'subheader', 'accent', 'all', 'center')
                elif value and value == '•':
                    self.apply_cell_style(cell, 'normal', border_type='all', alignment_type='center')
                elif value:
                    self.apply_cell_style(cell, 'normal', border_type='all', alignment_type='left_top')

                ws.row_dimensions[row_idx].height = 25

        # Set column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 30

    def generate_excel_file(self, output_filename='IBM_Terraform_Training_Course_Details.xlsx'):
        """Generate the complete Excel workbook."""
        try:
            print("Generating IBM Cloud Terraform Training Excel workbook...")

            # Create all worksheets
            self.create_course_overview_sheet()
            self.create_daily_syllabus_sheet()
            self.create_lab_sessions_sheet()
            self.create_sandbox_environment_sheet()
            self.create_delivery_options_sheet()
            self.create_prerequisites_sheet()

            # Save the workbook
            output_path = Path(output_filename)
            self.workbook.save(output_path)

            print(f"✅ Excel workbook successfully generated: {output_path.absolute()}")
            print(f"📊 Worksheets created: {len(self.workbook.sheetnames)}")
            print(f"📋 Worksheet names: {', '.join(self.workbook.sheetnames)}")

            return True

        except Exception as e:
            print(f"❌ Error generating Excel workbook: {str(e)}")
            return False

def main():
    """Main function to generate the Excel workbook."""
    print("🚀 Starting IBM Cloud Terraform Training Excel Generator")
    print("=" * 60)
    
    # Generate the Excel workbook
    generator = TerraformTrainingExcelGenerator()
    success = generator.generate_excel_file()
    
    if success:
        print("\n🎉 Excel generation completed successfully!")
        print("📄 The file is ready for client presentation.")
    else:
        print("\n❌ Excel generation failed. Please check the error messages above.")
        sys.exit(1)

if __name__ == "__main__":
    main()
